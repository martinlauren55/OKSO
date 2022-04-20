from openpyxl import load_workbook


def splitter(value):
    if len(value) == 6:
        f = value[0:2]
        s = value[2:4]
        e = value[4:6]
        res = f"{f}.{s}.{e}"
        return res
    else:
        pass


# Load in the workbook
wb = load_workbook('test.xlsm')
sheet = wb.active

# print(sheet.max_row)

for i in range(1, sheet.max_row + 1):
    c1 = sheet.cell(row=i, column=1).value
    c2 = sheet.cell(row=i, column=2).value
    c3 = sheet.cell(row=i, column=3).value
    # c4 = sheet.cell(row=i, column=4).value
    # res = f"{c1}@{c2}@{c3}@{c4}"
    res = f"{c1}@{c2}@{c3}"
    with open('1_from_exel.txt', "ab+") as f:
        f.write(bytes(res + '\n', 'utf8'))
